import io
import json
import base64

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError, NotFound
from drf_spectacular.utils import extend_schema
from .models import Proveedor, Insumo, FacturaCompra, MaterialCompra
from desarrollo.models import PedidoMaterial
from .serializers import ProveedorSerializer, InsumoSerializer, FacturaCompraSerializer
from desarrollo.serializers import PedidoMaterialSerializer
from django.db import transaction


class ComprasViewSet(viewsets.ViewSet):

    @extend_schema(responses={200: dict})
    @action(detail=False, methods=['get'], url_path='pedidos-material')
    def list_pedidos_pendientes(self, request):
        pedidos = PedidoMaterial.objects.filter(estado="generado").order_by('-created_at')
        return Response({"data": PedidoMaterialSerializer(pedidos, many=True).data})

    @extend_schema(request=None, responses={200: dict})
    @action(detail=False, methods=['post'], url_path='proveedores')
    def create_proveedor(self, request):
        proveedor = Proveedor.objects.create(nombre=request.data.get('nombre'))
        return Response({"message": "Proveedor creado", "data": {"id": proveedor.id, "nombre": proveedor.nombre}})

    @extend_schema(responses={200: dict})
    @action(detail=False, methods=['get'], url_path='proveedores')
    def list_proveedores(self, request):
        q = request.query_params.get('q')
        queryset = Proveedor.objects.all()
        if q:
            queryset = queryset.filter(nombre__icontains=q)
        return Response({"data": ProveedorSerializer(queryset, many=True).data})

    @extend_schema(request=None, responses={200: dict})
    @action(detail=False, methods=['post'], url_path='insumos')
    def create_insumo(self, request):
        nombre = request.data.get('nombre', '').strip()
        if not nombre:
            raise ValidationError("El nombre del artículo es obligatorio")
        insumo, created = Insumo.objects.update_or_create(
            nombre=nombre,
            defaults={
                'categoria': request.data.get('categoria', '').strip(),
                'subcategoria': request.data.get('subcategoria', '').strip(),
                'unidad': request.data.get('unidad', 'u').strip(),
            }
        )
        cantidad = request.data.get('cantidad')
        if cantidad is not None:
            try:
                cantidad_val = float(cantidad)
                if cantidad_val > 0:
                    from panol.models import Stock
                    Stock.objects.update_or_create(insumo=insumo, defaults={'cantidad': cantidad_val})
            except (ValueError, TypeError):
                pass
        msg = "Artículo creado" if created else "Artículo actualizado"
        return Response({"message": msg, "data": InsumoSerializer(insumo).data})

    @extend_schema(responses={200: dict})
    @action(detail=False, methods=['get'], url_path='insumos')
    def list_insumos(self, request):
        q = request.query_params.get('q')
        categoria = request.query_params.get('categoria')
        queryset = Insumo.objects.all().order_by('categoria', 'subcategoria', 'nombre')
        if q:
            queryset = queryset.filter(nombre__icontains=q)
        if categoria:
            queryset = queryset.filter(categoria__icontains=categoria)
        return Response({"data": InsumoSerializer(queryset, many=True).data})

    @extend_schema(responses={200: bytes})
    @action(detail=False, methods=['get'], url_path='insumos/template')
    def download_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Artículos"

        headers = ["Nombre (*)", "Categoría", "Subcategoría", "Unidad", "Cantidad Inicial"]
        header_fill = PatternFill(fill_type="solid", fgColor="1B2B3E")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            bottom=Side(style="thin", color="4A90D9"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        ws.row_dimensions[1].height = 22
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18

        ejemplos = [
            ["Tornillo M6 x 25mm", "Ferretería", "Tornillería", "u", 100],
            ["Cable 2.5mm THW", "Eléctrico", "Cables", "m", 50],
            ["Tuerca M6 hexagonal", "Ferretería", "Tornillería", "u", 200],
            ["Pintura epóxica gris", "Pintura", "Epóxica", "l", 10],
        ]
        example_font = Font(color="666666", italic=True)
        for row_data in ejemplos:
            ws.append(row_data)
            for col in range(1, 6):
                ws.cell(row=ws.max_row, column=col).font = example_font

        nota_row = ws.max_row + 2
        ws.cell(row=nota_row, column=1, value="(*) Nombre es obligatorio. Cantidad Inicial es opcional — si > 0 actualiza el stock en Pañol.")
        ws.cell(row=nota_row, column=1).font = Font(color="888888", italic=True, size=9)
        ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=5)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_articulos.xlsx"'
        return response

    @extend_schema(request=None, responses={200: dict})
    @action(detail=False, methods=['post'], url_path='insumos/bulk')
    @transaction.atomic
    def bulk_upload_articulos(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            raise ValidationError("Se requiere un archivo Excel (.xlsx)")

        try:
            wb = load_workbook(filename=io.BytesIO(archivo.read()), data_only=True)
            ws = wb.active
        except Exception:
            raise ValidationError("Archivo inválido. Usá la plantilla descargada desde el ERP.")

        creados = 0
        actualizados = 0
        errores = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell for cell in row if cell is not None):
                continue

            nombre = str(row[0]).strip() if row[0] is not None else ''
            if not nombre or nombre.startswith('(*)'):
                continue

            categoria = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            subcategoria = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            unidad = str(row[3]).strip() if len(row) > 3 and row[3] is not None else 'u'

            cantidad_raw = row[4] if len(row) > 4 and row[4] is not None else None

            try:
                insumo, created = Insumo.objects.update_or_create(
                    nombre=nombre,
                    defaults={'categoria': categoria, 'subcategoria': subcategoria, 'unidad': unidad},
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1
                if cantidad_raw is not None:
                    try:
                        cantidad_val = float(cantidad_raw)
                        if cantidad_val > 0:
                            from panol.models import Stock
                            Stock.objects.update_or_create(insumo=insumo, defaults={'cantidad': cantidad_val})
                    except (ValueError, TypeError):
                        pass
            except Exception as e:
                errores.append({"fila": row_idx, "nombre": nombre, "error": str(e)})

        return Response({
            "message": f"Procesado: {creados} creados, {actualizados} actualizados, {len(errores)} errores",
            "data": {"creados": creados, "actualizados": actualizados, "errores": errores},
        })

    @extend_schema(request=None, responses={200: dict})
    @action(detail=False, methods=['post'], url_path='facturas')
    @transaction.atomic
    def registrar_factura(self, request):
        pedido_id = request.data.get('pedido_material_id')

        try:
            pedido = PedidoMaterial.objects.get(id=pedido_id)
        except PedidoMaterial.DoesNotExist:
            raise NotFound(f"Pedido de material {pedido_id} no encontrado en Desarrollo")

        if pedido.estado == "facturado":
            raise ValidationError(f"Pedido {pedido_id} ya fue facturado")

        # FormData envía materiales como string JSON
        materiales_raw = request.data.get('materiales', '[]')
        materiales_data = json.loads(materiales_raw) if isinstance(materiales_raw, str) else materiales_raw

        # Codificar PDF en base64 si viene adjunto
        pdf_archivo = None
        pdf_file = request.FILES.get('pdf')
        if pdf_file:
            pdf_archivo = {
                'datos': base64.b64encode(pdf_file.read()).decode('utf-8'),
                'nombre': pdf_file.name,
            }

        insumo_cache = {}
        proveedor_cache = {}
        monto_total = 0.0

        for m in materiales_data:
            monto_total += float(m.get('cantidad', 0)) * float(m.get('precio_unitario', 0))

            nombre_insumo = m.get('nombre')
            if nombre_insumo not in insumo_cache:
                insumo_obj, _ = Insumo.objects.get_or_create(nombre=nombre_insumo)
                insumo_cache[nombre_insumo] = insumo_obj

            proveedor_nombre = m.get('proveedor', '')
            if proveedor_nombre and proveedor_nombre not in proveedor_cache:
                proveedor_obj, _ = Proveedor.objects.get_or_create(nombre=proveedor_nombre)
                proveedor_cache[proveedor_nombre] = proveedor_obj

        factura = FacturaCompra.objects.create(
            pedido_material_id=pedido,
            monto_total=monto_total,
            estado="registrada",
            pdf_archivo=pdf_archivo,
        )

        for m in materiales_data:
            proveedor_nombre = m.get('proveedor', '')
            MaterialCompra.objects.create(
                factura_id=factura,
                insumo=insumo_cache[m.get('nombre')],
                cantidad=m.get('cantidad'),
                precio_unitario=m.get('precio_unitario'),
                unidad_medida=m.get('unidad_medida', 'unidades'),
                proveedor=proveedor_cache.get(proveedor_nombre),
            )

        pedido.estado = "facturado"
        pedido.save()

        return Response({
            "message": "Factura registrada y vinculada al pedido de material",
            "data": FacturaCompraSerializer(factura).data
        })

    @extend_schema(responses={200: dict})
    def get_factura(self, request, pk=None):
        try:
            factura = FacturaCompra.objects.prefetch_related(
                'materiales__insumo', 'materiales__proveedor'
            ).get(id=pk)
        except FacturaCompra.DoesNotExist:
            raise NotFound(f"Factura {pk} no encontrada")
        return Response({"data": FacturaCompraSerializer(factura).data})

    @extend_schema(responses={200: dict})
    @action(detail=False, methods=['get'], url_path='facturas')
    def list_facturas(self, request):
        facturas = FacturaCompra.objects.prefetch_related('materiales__insumo', 'materiales__proveedor').all().order_by('-created_at')
        return Response({"data": FacturaCompraSerializer(facturas, many=True).data})
